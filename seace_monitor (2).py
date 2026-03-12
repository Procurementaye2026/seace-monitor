import requests
import pandas as pd
import smtplib
import os
import re
import ssl
import urllib3
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from requests.adapters import HTTPAdapter
from urllib3.util.ssl_ import create_urllib3_context
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ── Adaptador SSL especial para servidores SEACE ──
class SeaceSSLAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = create_urllib3_context()
        ctx.set_ciphers("DEFAULT:@SECLEVEL=1")
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        kwargs["ssl_context"] = ctx
        return super().init_poolmanager(*args, **kwargs)


# ==== CONFIGURACION ====
CORREO_REMITENTE  = os.getenv("CORREO_REMITENTE", "")
CORREO_CONTRASENA = os.getenv("CORREO_CONTRASENA", "")
CORREO_DESTINO    = os.getenv("CORREO_DESTINO", "")
REGION_ACTIVA     = os.getenv("REGION_ACTIVA", "LIMA")
MONTO_MINIMO      = int(os.getenv("MONTO_MINIMO", "0"))
ANIO_ACTUAL       = str(datetime.now().year)
URL_BUSCADOR      = "https://prodapp2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml"

RUBROS = {
    "Tecnologia": [
        "laptop", "computadora", "impresora", "monitor", "teclado",
        "tablet", "proyector", "servidor", "ups", "switch", "router",
        "scanner", "toner", "equipo informatico", "equipo de computo",
        "hardware", "suministro informatico"
    ],
    "Limpieza": [
        "limpieza", "desinfeccion", "aseo", "utiles de limpieza",
        "higiene", "saneamiento", "fumigacion"
    ],
    "Computo y TI": [
        "soporte tecnico", "mantenimiento informatico",
        "reparacion de equipos", "instalacion de software",
        "redes", "cableado", "tecnologia de la informacion", "software"
    ],
    "Ferreteria": [
        "ferreteria", "herramientas", "materiales de construccion",
        "pintura", "tuberias", "cables electricos",
        "tornillos", "taladro", "gasfiteria"
    ]
}


# ==== HELPERS ====
def limpiar_monto(valor):
    if not valor or str(valor).strip() in ["", "N/A", "S/N"]:
        return 0
    try:
        return float(re.sub(r"[^\d.]", "", str(valor).replace(",", ".")))
    except Exception:
        return 0


def crear_sesion():
    s = requests.Session()
    s.mount("https://", SeaceSSLAdapter())
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept-Language": "es-PE,es;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    })
    return s


def obtener_viewstate(session):
    """Obtiene ViewState y cookies de la pagina principal del buscador."""
    try:
        r = session.get(URL_BUSCADOR, timeout=20, verify=False)
        print("    GET buscador -> Status: " + str(r.status_code) + " | Bytes: " + str(len(r.text)))
        soup = BeautifulSoup(r.text, "html.parser")
        vs = soup.find("input", {"name": "javax.faces.ViewState"})
        if vs:
            val = vs.get("value", "")
            print("    ViewState OK (" + str(len(val)) + " chars)")
            return val
        print("    ViewState NO encontrado en la pagina")
        return ""
    except Exception as e:
        print("  [ERROR ViewState]: " + str(e))
        return ""


def extraer_html_de_jsf_ajax(texto_respuesta):
    """
    El buscador SEACE responde con XML JSF AJAX:
    <?xml ...?><partial-response><changes>
      <update id="..."><![CDATA[ HTML AQUI ]]></update>
    </changes></partial-response>
    Esta funcion extrae todo el HTML de los CDATA.
    """
    html_extraido = ""
    try:
        # Intentar parsear como XML para obtener CDATA
        root = ET.fromstring(texto_respuesta)
        for update in root.iter("update"):
            if update.text:
                html_extraido += update.text
        if html_extraido:
            return html_extraido
    except ET.ParseError:
        pass

    # Fallback: extraer CDATA con regex
    cdatas = re.findall(r"<!\[CDATA\[(.*?)\]\]>", texto_respuesta, re.DOTALL)
    if cdatas:
        return " ".join(cdatas)

    # Ultimo recurso: devolver el texto original
    return texto_respuesta


def parsear_filas_tabla(html_content):
    """
    Extrae filas de la tabla de resultados del SEACE.
    El SEACE usa PrimeFaces DataTable con clases ui-widget-content.
    """
    resultados = []
    soup = BeautifulSoup(html_content, "html.parser")

    # Estrategia 1: buscar filas PrimeFaces
    filas = soup.find_all("tr", class_=re.compile(r"ui-widget-content|ui-datatable-even|ui-datatable-odd"))

    # Estrategia 2: buscar cualquier fila con datos
    if not filas:
        tabla = soup.find("table", class_=re.compile(r"ui-datatable|resultados"))
        if tabla:
            filas = tabla.find_all("tr")[1:]  # Saltar encabezado

    # Estrategia 3: todas las filas con al menos 4 celdas
    if not filas:
        todas = soup.find_all("tr")
        filas = [f for f in todas if len(f.find_all("td")) >= 4]

    for fila in filas:
        celdas = fila.find_all("td")
        if len(celdas) >= 4:
            # Limpiar texto de cada celda
            textos = [c.get_text(separator=" ", strip=True) for c in celdas]
            resultados.append({
                "Entidad":       textos[0] if len(textos) > 0 else "N/A",
                "Descripcion":   textos[1] if len(textos) > 1 else "N/A",
                "Tipo Proceso":  textos[2] if len(textos) > 2 else "N/A",
                "Valor (S/.)":   textos[3] if len(textos) > 3 else "N/A",
                "Fecha Inicio":  textos[4] if len(textos) > 4 else "N/A",
                "Estado":        textos[5] if len(textos) > 5 else "N/A",
            })
    return resultados


def buscar_palabra(session, viewstate, palabra, debug_guardado):
    """Busca una palabra clave y retorna lista de resultados."""
    resultados = []
    try:
        payload = {
            "javax.faces.partial.ajax": "true",
            "javax.faces.source": "frmBuscarProceso:btnBuscar",
            "javax.faces.partial.execute": "@all",
            "javax.faces.partial.render": "frmBuscarProceso",
            "frmBuscarProceso:btnBuscar": "frmBuscarProceso:btnBuscar",
            "frmBuscarProceso": "frmBuscarProceso",
            "frmBuscarProceso:txtDescripcion": palabra,
            "frmBuscarProceso:ddlAnio": ANIO_ACTUAL,
            "frmBuscarProceso:ddlDpto": "15",
            "frmBuscarProceso:ddlTipoProceso": "",
            "javax.faces.ViewState": viewstate,
        }
        hdrs = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Faces-Request": "partial/ajax",
            "Accept": "application/xml, text/xml, */*; q=0.01",
            "Origin": "https://prodapp2.seace.gob.pe",
            "Referer": URL_BUSCADOR,
        }
        r = session.post(URL_BUSCADOR, data=payload,
                         headers=hdrs, timeout=20, verify=False)

        if r.status_code == 200 and len(r.text) > 100:
            # Guardar DEBUG solo para la primera palabra (para no llenar logs)
            if not debug_guardado[0]:
                debug_guardado[0] = True
                debug_path = "/tmp/debug_respuesta_seace.txt"
                with open(debug_path, "w", encoding="utf-8") as f:
                    f.write("=== STATUS: " + str(r.status_code) + " ===\n")
                    f.write("=== HEADERS: " + str(dict(r.headers)) + " ===\n\n")
                    f.write("=== RESPUESTA (primeros 3000 chars) ===\n")
                    f.write(r.text[:3000])
                print("    [DEBUG] Respuesta guardada en " + debug_path)

            # Extraer HTML del XML JSF AJAX
            html_content = extraer_html_de_jsf_ajax(r.text)

            # Parsear filas de la tabla
            filas_data = parsear_filas_tabla(html_content)
            resultados = filas_data

            if resultados:
                print("    '" + palabra + "': " + str(len(resultados)) + " resultados")
            else:
                # Mostrar primeros 200 chars del HTML extraido para debug
                print("    '" + palabra + "': 0 resultados | HTML(" + str(len(html_content)) + " bytes)")

    except Exception as e:
        print("  [ERROR '" + palabra + "']: " + str(e))

    return resultados


# ==== BUSQUEDA PRINCIPAL ====
def buscar_en_seace():
    sep = "=" * 55
    print(sep)
    print("   MONITOR SEACE - " + datetime.now().strftime("%d/%m/%Y %H:%M"))
    print("   Region: " + REGION_ACTIVA + " | Anio: " + ANIO_ACTUAL)
    print(sep)

    resultados = {rubro: [] for rubro in RUBROS}
    resultados["Todos los Rubros"] = []
    desc_vistos = set()
    debug_guardado = [False]  # Flag para guardar debug solo 1 vez

    session = crear_sesion()

    print("\n[1] Conectando al buscador SEACE...")
    viewstate = obtener_viewstate(session)

    for rubro, palabras in RUBROS.items():
        print("\n[2] Buscando rubro: " + rubro)
        for palabra in palabras:
            items = buscar_palabra(session, viewstate, palabra, debug_guardado)
            for item in items:
                desc = item.get("Descripcion", "")
                if desc and desc not in desc_vistos and len(desc) > 3:
                    desc_vistos.add(desc)
                    monto = limpiar_monto(item.get("Valor (S/.)", "0"))
                    if monto >= MONTO_MINIMO:
                        item["Rubro"] = rubro
                        item["Palabra Clave"] = palabra
                        item["Fuente"] = "SEACE Buscador Publico"
                        resultados[rubro].append(item)
                        resultados["Todos los Rubros"].append(item)

    print("\n" + sep)
    print("   RESUMEN POR RUBRO")
    print(sep)
    for rubro in RUBROS:
        cant = len(resultados[rubro])
        print("   " + rubro + ": " + str(cant) + " oportunidades")
    total = len(resultados["Todos los Rubros"])
    print("   TOTAL: " + str(total) + " oportunidades")
    print(sep)
    return resultados


# ==== EXCEL ====
def aplicar_estilo(ws, color):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=11)
    brd = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = brd
            c.alignment = Alignment(horizontal="left", vertical="center")
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 55)
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20


def guardar_excel(resultados):
    nombre = "SEACE_Oportunidades_" + datetime.now().strftime("%Y-%m-%d") + ".xlsx"
    ruta = "/tmp/" + nombre
    colores = {
        "Resumen":          "1a73e8",
        "Tecnologia":       "0f9d58",
        "Limpieza":         "f4b400",
        "Computo y TI":     "4285f4",
        "Ferreteria":       "db4437",
        "Todos los Rubros": "37474f"
    }
    with pd.ExcelWriter(ruta, engine="openpyxl") as w:
        resumen = []
        for r, l in resultados.items():
            if r != "Todos los Rubros":
                resumen.append({
                    "Rubro":         r,
                    "Oportunidades": len(l),
                    "Fecha":         datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "Region":        REGION_ACTIVA
                })
        pd.DataFrame(resumen).to_excel(w, sheet_name="Resumen", index=False)
        for rubro, lista in resultados.items():
            nombre_hoja = rubro[:31]
            if lista:
                df = pd.DataFrame(lista).drop_duplicates(subset=["Descripcion"])
            else:
                df = pd.DataFrame([{"Mensaje": "Sin resultados para " + rubro}])
            df.to_excel(w, sheet_name=nombre_hoja, index=False)
    wb = load_workbook(ruta)
    for s in wb.sheetnames:
        color = next((v for k, v in colores.items() if k in s), "1a73e8")
        aplicar_estilo(wb[s], color)
    wb.save(ruta)
    print("\n[Excel] Guardado: " + ruta)
    return ruta


# ==== CORREO ====
def enviar_correo(resultados, ruta_excel):
    total = len(resultados.get("Todos los Rubros", []))
    fecha = datetime.now().strftime("%d/%m/%Y")

    colores_bg = {
        "Tecnologia":   "#e8f5e9",
        "Limpieza":     "#fff9c4",
        "Computo y TI": "#e3f2fd",
        "Ferreteria":   "#fce4ec"
    }

    filas_list = []
    for r, l in resultados.items():
        if r == "Todos los Rubros":
            continue
        bg = colores_bg.get(r, "#f5f5f5")
        fila = (
            "<tr style='background:" + bg + "'>"
            "<td style='padding:8px'>" + r + "</td>"
            "<td style='padding:8px;text-align:center;font-size:18px'>"
            "<b>" + str(len(l)) + "</b></td></tr>"
        )
        filas_list.append(fila)
    filas = "".join(filas_list)

    top_list = []
    for o in resultados.get("Todos los Rubros", [])[:8]:
        entidad = str(o.get("Entidad", ""))[:40]
        desc    = str(o.get("Descripcion", ""))[:55]
        monto   = str(o.get("Valor (S/.)", ""))
        rubro   = str(o.get("Rubro", ""))
        fila = (
            "<tr>"
            "<td style='padding:6px'>" + entidad + "</td>"
            "<td style='padding:6px'>" + desc + "...</td>"
            "<td style='padding:6px'>" + monto + "</td>"
            "<td style='padding:6px'>" + rubro + "</td>"
            "</tr>"
        )
        top_list.append(fila)

    if top_list:
        top_ops = "".join(top_list)
    else:
        top_ops = (
            "<tr><td colspan='4' style='padding:15px;text-align:center'>"
            "Sin oportunidades hoy</td></tr>"
        )

    html = (
        "<html><body style='font-family:Arial,sans-serif;background:#f5f5f5;padding:20px'>"
        "<div style='max-width:780px;margin:auto;background:white;border-radius:8px'>"
        "<div style='background:#1a73e8;padding:25px;color:white;text-align:center'>"
        "<h1 style='margin:0'>Monitor SEACE - Lima</h1>"
        "<p style='margin:8px 0 0'>Reporte Diario - " + fecha + "</p>"
        "</div><div style='padding:20px'>"
        "<h2 style='color:#1a73e8'>Resumen por Rubro</h2>"
        "<table style='width:100%;border-collapse:collapse'>"
        "<tr style='background:#1a73e8;color:white'>"
        "<th style='padding:10px'>Rubro</th>"
        "<th style='padding:10px'>Oportunidades</th></tr>"
        + filas
        + "<tr style='background:#e8eaf6'>"
        "<td style='padding:10px'><b>TOTAL</b></td>"
        "<td style='padding:10px;text-align:center;font-size:20px;color:#1a73e8'>"
        "<b>" + str(total) + "</b></td></tr></table>"
        "<h2 style='color:#1a73e8;margin-top:25px'>Top 8 Oportunidades</h2>"
        "<table style='width:100%;border-collapse:collapse;font-size:13px'>"
        "<tr style='background:#1a73e8;color:white'>"
        "<th style='padding:8px'>Entidad</th>"
        "<th style='padding:8px'>Descripcion</th>"
        "<th style='padding:8px'>Monto</th>"
        "<th style='padding:8px'>Rubro</th></tr>"
        + top_ops
        + "</table>"
        "<p style='color:#666;font-size:12px;margin-top:20px'>"
        "Excel adjunto con detalle completo</p>"
        "</div></div></body></html>"
    )

    msg = MIMEMultipart("alternative")
    msg["Subject"] = "SEACE Lima - " + str(total) + " Oportunidades (" + fecha + ")"
    msg["From"]    = CORREO_REMITENTE
    msg["To"]      = CORREO_DESTINO
    msg.attach(MIMEText(html, "html"))

    if ruta_excel and os.path.exists(ruta_excel):
        with open(ruta_excel, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                "attachment; filename=" + os.path.basename(ruta_excel)
            )
            msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(CORREO_REMITENTE, CORREO_CONTRASENA)
            s.sendmail(CORREO_REMITENTE, CORREO_DESTINO, msg.as_string())
        print("[Correo] Enviado exitosamente a " + CORREO_DESTINO)
    except Exception as e:
        print("[Correo] Error al enviar: " + str(e))


# ==== MAIN ====
if __name__ == "__main__":
    datos   = buscar_en_seace()
    archivo = guardar_excel(datos)
    enviar_correo(datos, archivo)
    print("\n[FIN] Ejecucion completada.")
