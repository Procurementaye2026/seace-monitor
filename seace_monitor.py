import requests, pandas as pd, smtplib, os, re
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import ssl
import urllib3
from requests.adapters import HTTPAdapter
from urllib3.util.ssl_ import create_urllib3_context

# Silenciar advertencias SSL (el servidor SEACE usa certificado antiguo)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Adaptador especial para servidores con DH key pequeña (SEACE)
class SeaceSSLAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = create_urllib3_context()
        ctx.set_ciphers("DEFAULT:@SECLEVEL=1")
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        kwargs["ssl_context"] = ctx
        return super().init_poolmanager(*args, **kwargs)


# ==== CONFIGURACION ====
CORREO_REMITENTE    = os.getenv("CORREO_REMITENTE", "")
CORREO_CONTRASENA   = os.getenv("CORREO_CONTRASENA", "")
CORREO_DESTINO      = os.getenv("CORREO_DESTINO", "")
REGION_ACTIVA       = os.getenv("REGION_ACTIVA", "LIMA")
MONTO_MINIMO        = int(os.getenv("MONTO_MINIMO", "0"))

# Palabras clave por rubro
RUBROS = {
    "Tecnologia": [
        "laptop", "computadora", "pc ", "impresora", "monitor",
        "teclado", "mouse", "disco duro", "memoria ram", "tablet",
        "proyector", "servidor", "ups", "switch", "router", "scanner",
        "toner", "equipo informatico", "equipo de computo",
        "equipo tecnologico", "hardware", "suministro informatico"
    ],
    "Limpieza": [
        "limpieza", "desinfeccion", "aseo", "utiles de limpieza",
        "higiene", "saneamiento", "fumigacion", "desratizacion"
    ],
    "Computo y TI": [
        "mantenimiento de computo", "soporte tecnico",
        "mantenimiento informatico", "reparacion de equipos",
        "instalacion de software", "redes", "cableado",
        "soporte de sistemas", "tecnologia de la informacion",
        "servicios de tecnologia", "software", "infraestructura ti"
    ],
    "Ferreteria": [
        "ferreteria", "herramientas", "materiales de construccion",
        "pintura", "tuberias", "cables electricos", "candado",
        "cerradura", "tornillos", "taladro", "soldadura",
        "material ferretero", "gasfiteria", "plomeria"
    ]
}

ANIO_ACTUAL = str(datetime.now().year)
URL_BUSCADOR = "https://prodapp2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/buscadorPublico.xhtml"

# ==== HELPERS ====
def limpiar_monto(valor):
    if not valor or str(valor).strip() in ["", "N/A", "S/N"]: return 0
    try:
        return float(re.sub(r"[^\d.]", "", str(valor).replace(",", ".")))
    except:
        return 0

def obtener_viewstate(session):
    """Obtiene cookies y ViewState del buscador JSF"""
    try:
         r = session.get(URL_BUSCADOR, timeout=20, verify=False,
                        headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
        soup = BeautifulSoup(r.text, "html.parser")
        vs = soup.find("input", {"name": "javax.faces.ViewState"})
        return vs["value"] if vs else ""
    except Exception as e:
        print(f"  [ERROR ViewState]: {e}")
        return ""

def buscar_palabra(session, viewstate, palabra, codigo_dpto="15"):
    """Busca una palabra clave en el buscador público JSF"""
    resultados = []
    try:
        payload = {
            "javax.faces.partial.ajax": "true",
            "javax.faces.source": "frmBuscarProceso:btnBuscar",
            "javax.faces.partial.execute": "@all",
            "javax.faces.partial.render": "frmBuscarProceso",
            "frmBuscarProceso:btnBuscar": "frmBuscarProceso:btnBuscar",
            "frmBuscarProceso": "frmBuscarProceso",
            "frmBuscarProceso:txtDescripcion": palabra,
            "frmBuscarProceso:ddlAnio": ANIO_ACTUAL,
            "frmBuscarProceso:ddlDpto": codigo_dpto,
            "frmBuscarProceso:ddlTipoProceso": "",
            "javax.faces.ViewState": viewstate,
        }
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "X-Requested-With": "XMLHttpRequest",
            "Faces-Request": "partial/ajax",
            "Origin": "https://prodapp2.seace.gob.pe",
            "Referer": URL_BUSCADOR,
        }
       r = session.post(URL_BUSCADOR, data=payload, headers=headers, timeout=15, verify=False)
        if r.status_code == 200 and len(r.text) > 100:
            soup = BeautifulSoup(r.text, "html.parser")
            filas = soup.find_all("tr", class_=re.compile(r"ui-widget-content|fila"))
            if not filas:
                filas = soup.find_all("tr")[1:]  # saltar encabezado
            for fila in filas:
                celdas = fila.find_all("td")
                if len(celdas) >= 4:
                    resultados.append({
                        "Entidad":      celdas[0].get_text(strip=True) if len(celdas) > 0 else "N/A",
                        "Descripcion":  celdas[1].get_text(strip=True) if len(celdas) > 1 else "N/A",
                        "Tipo Proceso": celdas[2].get_text(strip=True) if len(celdas) > 2 else "N/A",
                        "Valor (S/.)":  celdas[3].get_text(strip=True) if len(celdas) > 3 else "N/A",
                        "Fecha Inicio": celdas[4].get_text(strip=True) if len(celdas) > 4 else "N/A",
                        "Estado":       celdas[5].get_text(strip=True) if len(celdas) > 5 else "N/A",
                        "Palabra Clave": palabra,
                        "Fuente":       "SEACE Buscador Publico"
                    })
    except Exception as e:
        print(f"  [ERROR busqueda '{palabra}']: {e}")
    return resultados

# ==== BUSQUEDA PRINCIPAL ====
def buscar_en_seace():
    print(f"\n{'='*55}")
    print(f"   MONITOR SEACE - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"   Region: {REGION_ACTIVA} | Anio: {ANIO_ACTUAL}")
    print(f"{'='*55}")

    resultados = {rubro: [] for rubro in RUBROS}
    resultados["Todos los Rubros"] = []
    desc_vistos = set()

    session = requests.Session()
    session.mount("https://", SeaceSSLAdapter())
    session.headers.update({"Accept-Language": "es-PE,es;q=0.9"})9"})

    print("\n[1] Obteniendo sesion del buscador SEACE...")
    viewstate = obtener_viewstate(session)
    if viewstate:
        print(f"    ViewState OK ({len(viewstate)} chars)")
    else:
        print("    ADVERTENCIA: No se obtuvo ViewState, intentando sin el...")

    # Buscar cada palabra clave
    for rubro, palabras in RUBROS.items():
        print(f"\n[2] Buscando rubro: {rubro}")
        for palabra in palabras:
            items = buscar_palabra(session, viewstate, palabra)
            for item in items:
                desc = item.get("Descripcion", "")
                if desc not in desc_vistos:
                    desc_vistos.add(desc)
                    item["Rubro"] = rubro
                    monto = limpiar_monto(item.get("Valor (S/.)","0"))
                    if monto >= MONTO_MINIMO:
                        resultados[rubro].append(item)
                        resultados["Todos los Rubros"].append(item)
            if items:
                print(f"    '{palabra}': {len(items)} resultados")

    # Resumen
    print(f"\n{'='*55}")
    print("   RESUMEN POR RUBRO")
    print(f"{'='*55}")
    iconos = {"Tecnologia":"💻","Limpieza":"🧹","Computo y TI":"🖥️","Ferreteria":"🔧"}
    for rubro, lista in resultados.items():
        if rubro != "Todos los Rubros":
            print(f"   {iconos.get(rubro,'📁')} {rubro:<20}: {len(lista)} oportunidades")
    print(f"   📊 {'TOTAL':<22}: {len(resultados['Todos los Rubros'])} oportunidades")
    print(f"{'='*55}")
    return resultados

# ==== EXCEL ====
def aplicar_estilo(ws, color):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=11)
    brd  = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    for c in ws[1]:
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = brd
            c.alignment = Alignment(horizontal="left", vertical="center")
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+4, 55)
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20

def guardar_excel(resultados):
    ruta = f"/tmp/SEACE_Oportunidades_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    colores = {"Resumen":"1a73e8","Tecnologia":"0f9d58",
               "Limpieza":"f4b400","Computo y TI":"4285f4","Ferreteria":"db4437"}
    with pd.ExcelWriter(ruta, engine="openpyxl") as w:
        resumen = [{"Rubro": r, "Oportunidades": len(l),
                    "Fecha": datetime.now().strftime("%d/%m/%Y %H:%M"), "Region": REGION_ACTIVA}
                   for r, l in resultados.items() if r != "Todos los Rubros"]
        pd.DataFrame(resumen).to_excel(w, sheet_name="Resumen", index=False)
        for rubro, lista in resultados.items():
            nombre = rubro[:31]
            df = pd.DataFrame(lista) if lista else pd.DataFrame([{"Mensaje": f"Sin resultados para {rubro}"}])
            df.to_excel(w, sheet_name=nombre, index=False)
    wb = load_workbook(ruta)
    for s in wb.sheetnames:
        color = next((v for k,v in colores.items() if k in s), "1a73e8")
        aplicar_estilo(wb[s], color)
    wb.save(ruta)
    print(f"\n[Excel] Guardado: {ruta}")
    return ruta

# ==== CORREO ====
def enviar_correo(resultados, ruta_excel):
    total = len(resultados.get("Todos los Rubros", []))
    fecha = datetime.now().strftime("%d/%m/%Y")
    iconos     = {"Tecnologia": "💻", "Limpieza": "🧹",
                  "Computo y TI": "🖥", "Ferreteria": "🔧"}
    colores_bg = {"Tecnologia": "#e8f5e9", "Limpieza": "#fff9c4",
                  "Computo y TI": "#e3f2fd", "Ferreteria": "#fce4ec"}
    color_def  = "#f5f5f5"
    icono_def  = "📁"

    # ── Filas resumen ──
    filas_list = []
    for r, l in resultados.items():
        if r == "Todos los Rubros":
            continue
        bg  = colores_bg.get(r, color_def)
        ico = iconos.get(r, icono_def)
        filas_list.append(
            f"<tr style='background:{bg}'>"
            f"<td style='padding:8px'>{ico} {r}</td>"
            f"<td style='padding:8px;text-align:center;font-size:18px'><b>{len(l)}</b></td>"
            f"</tr>"
        )
    filas = "".join(filas_list)

    # ── Top 8 oportunidades ──
    top_list = []
    for o in resultados.get("Todos los Rubros", [])[:8]:
        entidad = o.get("Entidad", "")[:40]
        desc    = o.get("Descripcion", "")[:55]
        monto   = o.get("Valor (S/.)", "")
        rubro   = o.get("Rubro", "")
        top_list.append(
            f"<tr>"
            f"<td style='padding:6px'>{entidad}</td>"
            f"<td style='padding:6px'>{desc}...</td>"
            f"<td style='padding:6px'>{monto}</td>"
            f"<td style='padding:6px'>{rubro}</td>"
            f"</tr>"
        )
    top_ops = "".join(top_list) if top_list else (
        "<tr><td colspan='4' style='padding:15px;text-align:center'>"
        "Sin oportunidades hoy</td></tr>"
    )

    html = (
        "<html><body style='font-family:Arial,sans-serif;background:#f5f5f5;padding:20px'>"
        "<div style='max-width:780px;margin:auto;background:white;border-radius:8px;overflow:hidden'>"
        "<div style='background:linear-gradient(135deg,#1a73e8,#0d47a1);padding:28px;"
        "color:white;text-align:center'>"
        "<h1 style='margin:0'>🏛 Monitor SEACE - Lima</h1>"
        f"<p style='margin:8px 0 0'>Reporte Diario - {fecha}</p>"
        "</div><div style='padding:20px'>"
        "<h2 style='color:#1a73e8'>📊 Resumen por Rubro</h2>"
        "<table style='width:100%;border-collapse:collapse'>"
        "<tr style='background:#1a73e8;color:white'>"
        "<th style='padding:10px'>Rubro</th>"
        "<th style='padding:10px'>Oportunidades</th></tr>"
        + filas +
        "<tr style='background:#e8eaf6'>"
        "<td style='padding:10px'><b>📊 TOTAL</b></td>"
        f"<td style='padding:10px;text-align:center;font-size:20px;color:#1a73e8'><b>{total}</b></td>"
        "</tr></table>"
        "<h2 style='color:#1a73e8;margin-top:25px'>🔝 Top 8 Oportunidades</h2>"
        "<table style='width:100%;border-collapse:collapse;font-size:13px'>"
        "<tr style='background:#1a73e8;color:white'>"
        "<th style='padding:8px'>Entidad</th>"
        "<th style='padding:8px'>Descripcion</th>"
        "<th style='padding:8px'>Monto</th>"
        "<th style='padding:8px'>Rubro</th></tr>"
        + top_ops +
        "</table>"
        "<p style='color:#666;font-size:12px;margin-top:20px'>"
        "📎 Excel adjunto con detalle completo<br>"
        "🔗 <a href='https://prodapp2.seace.gob.pe/seacebus-uiwd-pub/buscadorPublico/"
        "buscadorPublico.xhtml'>Ir al SEACE</a></p>"
        "</div></div></body></html>"
    )

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"SEACE Lima - {total} Oportunidades ({fecha})"
    msg["From"]    = CORREO_REMITENTE
    msg["To"]      = CORREO_DESTINO
    msg.attach(MIMEText(html, "html"))

    if ruta_excel and os.path.exists(ruta_excel):
        with open(ruta_excel, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={os.path.basename(ruta_excel)}"
            )
            msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(CORREO_REMITENTE, CORREO_CONTRASENA)
            s.sendmail(CORREO_REMITENTE, CORREO_DESTINO, msg.as_string())
        print(f"[Correo] Enviado a {CORREO_DESTINO}")
    except Exception as e:
        print(f"[Correo] Error: {e}")
# ==== MAIN ====
if __name__ == "__main__":
    datos   = buscar_en_seace()
    archivo = guardar_excel(datos)
    enviar_correo(datos, archivo)
    print("\n[FIN] Ejecucion completada.")


